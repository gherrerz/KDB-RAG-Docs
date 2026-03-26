"""XLSX parser for extracting cell values."""

from __future__ import annotations

from pathlib import Path


def parse_xlsx(file_path: Path) -> str:
    """Extract worksheet values from an XLSX file.

    Returns an empty string when parsing fails so ingestion can continue.
    """
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(
            filename=str(file_path),
            read_only=True,
            data_only=True,
        )
        lines: list[str] = []

        for sheet in workbook.worksheets:
            lines.append(f"[Sheet] {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                values = [str(value).strip() for value in row if value is not None]
                if values:
                    lines.append("\t".join(values))

        return "\n".join(lines)
    except Exception:
        return ""
